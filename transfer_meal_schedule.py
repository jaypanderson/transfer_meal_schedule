import openpyxl
import os
import sys
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.chart import BarChart, Reference
from tkinter import filedialog
from typing import Union
from copy import copy
from itertools import zip_longest


def choose_file(file_type: int) -> str:
    if file_type == 1:
        title = '献立表を選択してください。'
    elif file_type == 2:
        title = '離乳食献立を選択してください'
    elif file_type == 3:
        title = '検食簿原本を選択してください。'
    else:
        title = ''
    return filedialog.askopenfilename(title=title)


def find_start_of_dates(sheet: Worksheet) -> int:
    """
    This finds the row that contains the date (int in this case) which
    :param sheet:
    :return:
    """
    for i, row in enumerate(sheet.iter_rows(), start=1):
        if isinstance(row[0].value, int):
            # we add one because openpyxl uses 1 indexing.
            return i


def find_date_ranges(sheet: Worksheet) -> dict[int, tuple[str, int, int]]:
    """
    Finds how many rows each meal uses and saves that information into a dictionary for later use. In the Excel file the
    meal schedule for a particular date is spread into multiple rows for readability. This functions finds the start and
    end row for all dates.
    :param sheet: The sheet where the ranges will be gathered from.
    :return: A dictionary that contains the date as the keys and the row ranges as the value that represent which rows
    are part of the meal schedule for a given date. The dictionary is formatted as the following.
    {date: (day, start, end), ...}, ex) {2: ('木', 6, 10)}. The ranges are formatted to be inclusive since openpyxl
    typically uses inclusive ranges.
    """
    start_row = find_start_of_dates(sheet)
    date_ranges = {}
    date = None
    day = None
    start = None
    # TODO this code can probably use some refactoring to make it cleaner
    for i, row in enumerate(sheet.iter_rows(min_row=start_row), start=start_row):
        if i == start_row:
            start = i
            date = row[0].value
            day = row[1].value
        elif row[0].value is not None:
            end = i - 1
            date_ranges[date] = (day, start, end)
            start = i
            date = row[0].value
            day = row[1].value

    return date_ranges


def gather_text_big_kids(sheet: Worksheet, start: int, end: int) -> tuple[str]:
    """
    A helper function to find text from specific cells in the worksheet and bundle it.
    :param sheet: The sheet where the text will be gathered from.
    :param start: The start of the row range of where we are looking for the text.
    :param end: The end of the row range of where we are looking for the text
    :return: A packaged tuple of all the text data that will be inserted into a dictionary for later use.
    """
    breakfast = []
    lunch = []
    snack = []
    for row in sheet.iter_rows(min_row=start, max_row=end):
        if row[2].value:
            lunch.append(row[2].value)
        if row[6].value:
            breakfast.append(row[6].value)
        if row[7].value:
            snack.append(row[7].value)
    return '\n'.join(breakfast), '\n'.join(lunch), '\n'.join(snack)


def extract_meal_data_big_kids(path: str) -> dict:
    """
    A function that finds how many days the meal schedule has as well as finding how many rows in the sheet each meal of
    that day takes up, and places each meal for that day in an organized.dictionary.
    :param path: The file path for the meal schedule.
    :return: A dictionary of the meal schedule organized by date. The format of the dictionary is as follows,
    {date: (day of the week, breakfast, lunch, snack), ...} and here is an example of the dictionary
    {28: ('木', '野菜ハイハイン\n\n', '', '', '５倍粥\n鶏ササミと野菜（人参・グリンピース）煮物\n玉ねぎとわかめの煮物',
    'さつま芋きなこがけ\n\n'), ...} In this example the early and middle meals are not there, so they are empty strings.
    """
    book = openpyxl.load_workbook(path)
    sheet = book.active
    date_ranges = find_date_ranges(sheet)
    meal_data_big_kids = {}
    for key, val in date_ranges.items():
        day = val[0]
        start = val[1]
        end = val[2]
        meal_data_big_kids[key] = (day,) + gather_text_big_kids(sheet, start, end)
    return meal_data_big_kids


def gather_text_small_kids(sheet: Worksheet, start: int, end: int, max_column: str) -> tuple[str]:
    """
    A helper function to find text from specific cells in the worksheet and bundle it.  Depending on the max_column
    the locations of the cells are slightly different.  This is due to the fact that depending on how far along we are
    in the year early middle and late meals may not be served.
    :param sheet: The sheet where the text will be gathered from.
    :param start: The start of the row range of where we are looking for the text.
    :param end: The end of the row range of where we are looking for the text
    :param max_column: The highest letter column being used.  if G, early middle and late are being served. if F only
    middle and late are being served and then if E only late is being served.
    :return: a packaged tuple of all the text data that will be inserted into a dictionary for later use.
    """
    breakfast, early, middle, late, snack = [], [], [], [], []
    meals = {'G': [early, middle, late, breakfast, snack], 'F': [middle, late, breakfast, snack],
             'E': [late, breakfast, snack]}

    for row in sheet.iter_rows(min_row=start, max_row=end):
        for i, meal in enumerate(meals[max_column], start=2):
            if row[i].value:
                meal.append(row[i].value)

    return '\n'.join(breakfast), '\n'.join(early), '\n'.join(middle), '\n'.join(late), '\n'.join(snack)


# TODO the bellow function is commented out due to the function above replacing it. Until further tests are done this
# TODO will remain here in case the above code is somehow incorrect.

# def __gather_text_small_kids(sheet: Worksheet, start: int, end: int, max_column: str) -> tuple[str]:
#     """
#     A helper function to find text from specific cells in the worksheet and bundle it.  Depending on the max_column
#     the locations of the cells are slightly different.  This is due to the fact that depending on how far along we are
#     in the year early middle and late meals may not be served.
#     :param sheet: The sheet where the text will be gathered from.
#     :param start: The start of the row range of where we are looking for the text.
#     :param end: The end of the row range of where we are looking for the text
#     :param max_column: The highest letter column being used.  if G, early middle and late are being served. if F only
#     middle and late are being served and then if E only late is being served.
#     :return: a packaged tuple of all the text data that will be inserted into a dictionary for later use.
#     """
#     breakfast = []
#     early = []
#     middle = []
#     late = []
#     snack = []
#     if max_column == 'G':
#         for row in sheet.iter_rows(min_row=start, max_row=end):
#             if row[2].value is not None:
#                 early.append(row[2].value)
#             if row[3].value is not None:
#                 middle.append(row[3].value)
#             if row[4].value is not None:
#                 late.append(row[4].value)
#             if row[5].value is not None:
#                 breakfast.append(row[5].value)
#             if row[6].value is not None:
#                 snack.append(row[6].value)
#     elif max_column == 'F':
#         for row in sheet.iter_rows(min_row=start, max_row=end):
#             if row[2].value is not None:
#                 middle.append(row[2].value)
#             if row[3].value is not None:
#                 late.append(row[3].value)
#             if row[4].value is not None:
#                 breakfast.append(row[4].value)
#             if row[5].value is not None:
#                 snack.append(row[5].value)
#     elif max_column == 'E':
#         for row in sheet.iter_rows(min_row=start, max_row=end):
#             if row[2].value is not None:
#                 late.append(row[2].value)
#             if row[3].value is not None:
#                 breakfast.append(row[3].value)
#             if row[4].value is not None:
#                 snack.append(row[4].value)
#
#     return '\n'.join(breakfast), '\n'.join(early), '\n'.join(middle), '\n'.join(late), '\n'.join(snack)


def extract_meal_data_small_kids(path: str) -> Union[dict, None]:
    """
    A function that finds how many days the meal schedule has as well as finding how many rows in the sheet each meal of
    that day takes up, and places each meal for that day in an organized.dictionary.
    :param path: The file path for the meal schedule.
    :return: A dictionary of the meal schedule organized by date The format of the dictionary is as follows
    {date: (day of the week, breakfast, early, middle, late, snack), ...} and here is an example of the dictionary
    {4: ('木', '●/▲たまごボーロ\n【リッツクラッカー】\nお茶', '新春ちらし寿司\n切干大根の煮物\n花麩のすまし汁', '上用まんじゅう\n▲牛乳【お茶】'), ...}
    the return can potentially be None if the path of the meal schedule is a empty string ('').  This happens when the
    user is prompted to choose a file and closes the window without choosing. This returns a '' empty string.  This will
    typically happen between the months of january and april where there is no meal schedule for the small kids because
    they start eating what is on the big kids meal schedule.
    """
    if path == '':
        return None
    book = openpyxl.load_workbook(path)
    sheet = book.active
    meal_data_small_kids = {}
    date_ranges = find_date_ranges(sheet)
    max_column = get_column_letter(sheet.max_column)
    for key, val in date_ranges.items():
        day = val[0]
        start = val[1]
        end = val[2]
        meal_data_small_kids[key] = (day,) + gather_text_small_kids(sheet, start, end, max_column)
    return meal_data_small_kids


# add result to the end of the file name
def new_file_path(path: str, added_text: str = 'result') -> str:
    """
    This function creates a new name for the path of a save file. This is to avoid saving over the original Excel file
    that was used to create the new one. It places a new text between the name and the extension name. If no added_text
    is provided the default 'result' will be used.
    :param path: The path of the original Excel file.
    :param added_text: The text that will be added inbetween the name and the extension name of the original path.
    :return: The newly formed name path where the new Excel file will be saved to.
    """
    idx = path.find('.')
    ans = path[:idx] + added_text + path[idx:]
    return ans


def copy_sheet(sheet: Worksheet, new_sheet: Worksheet) -> None:
    """
    copy the contents from one sheet to another sheet. This function only copies the contents of the cells and its
    style, other aspects of the sheet are copied with other functions.
    :param sheet:The base sheet from which the cells are copied.
    :param new_sheet: The new sheet where the cell contents and style are pasted.
    :return: None
    """
    for row in sheet:
        for cell in row:
            new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)

            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
                new_cell.comment = copy(cell.comment)


# replicate the cell merges from base template.
def merge_cells(sheet: Worksheet, new_sheet: Worksheet) -> None:
    """
    When a new sheet is created and everything is copied, the one thing that will not be copied is the location of where
    there are merged cells.  This function iterates over all the merged cells of the base worksheet and then merges
    those cells in the new sheet.
    :param sheet: The base work sheet this function will use to iterate over the location of the merged cells.
    :param new_sheet: The new work sheet where the function will merge the cells from the base work sheet.
    :return: None
    """
    for merged_cell_range in sheet.merged_cells.ranges:
        new_sheet.merge_cells(str(merged_cell_range))


# copy the width and height of cells
def copy_dimensions(sheet: Worksheet, new_sheet: Worksheet) -> None:
    """
    A function that changes the new sheet's size of the cells to match that of a another sheet. Currently, it is set up
    so that it only copies the width of the cells.  If the height also needs to be copied, then Uncomment that portion
    of the code and vice versa.
    :param sheet: The sheet from which the cell dimensions will be copied from
    :param new_sheet: The sheet where the dimensions will be pasted into.
    :return: None
    """
    for row, col in zip_longest(sheet.row_dimensions, sheet.column_dimensions):
        if row is not None:
            new_sheet.row_dimensions[row].height = sheet.row_dimensions[row].height
        if col is not None:
            new_sheet.column_dimensions[col].width = sheet.column_dimensions[col].width


# copy the part of the worksheet that will be printed onto the new sheet.
def copy_print_area(sheet: Worksheet, new_sheet: Worksheet) -> None:
    """
    A certain area of a page is selected as the default area to be printed when the print button is pressed. To avoid
    having to change the print area for every single new sheet this function copies the print area attribute from the
    base sheet to the new sheet.
    :param sheet: the base sheet from which we will copy the print area attribute.
    :param new_sheet: The new sheet where we will paste the print area.
    :return: None
    """
    if sheet.print_area:
        new_sheet.print_area = sheet.print_area


def copy_margins(sheet: Worksheet, new_sheet: Worksheet) -> None:
    """
    copy the peper margins from one sheet to another.
    :param sheet: original sheet the attributes will be copied from.
    :param new_sheet: the new sheet where the attributes will be pasted into.
    :return: None
    """
    new_sheet.page_margins = sheet.page_margins


def copy_page_size(sheet: Worksheet, new_sheet: Worksheet) -> None:
    """
    copy the paper size of the sheet so that when printing it knows which size to use.
    :param sheet: original sheet the attributes will be copied from.
    :param new_sheet: the new sheet where the attributes will be pasted into.
    :return: None
    """
    new_sheet.page_setup.paperSize = sheet.page_setup.paperSize


def copy_all_elements(sheet: Worksheet, new_sheet: Worksheet) -> None:
    """
    A utility function to run other smaller functions. In openpyexl there is no straight forward way to completely copy
    all aspects of a worksheet into another.  This requires accessing specific attributes that need to be copied.  In
    the future if there are new attributes that need to be copied, create and add another function to this list.
    :param sheet: original sheet the attributes will be copied from.
    :param new_sheet: the new sheet where the attributes will be pasted into.
    :return: None
    """
    copy_sheet(sheet, new_sheet)
    merge_cells(sheet, new_sheet)
    copy_dimensions(sheet, new_sheet)
    copy_print_area(sheet, new_sheet)
    copy_margins(sheet, new_sheet)
    copy_page_size(sheet, new_sheet)


# this function is so that when using the script the image retrieval works as well as when the executable
# is created as well so that I don't have to write different code from when im developing and when
# im deploying.
def resource_path(relative_path) -> str | bytes:
    """
    This function changes the path depending on if it accesses the image while in a .py file or in an executable file.
    Because when making an executable the location of the image attached may change from when accessing it from when
    accessing it while in a .py file, the whole path needs to be added and not just the relative path.  So essentially
    this function allows for the code to remain the same whether accessing the image normally or with an executable.
    Do not fully understand how it works, so do not delete.
    :param relative_path: the relative path for the object being accessed.
    :return: The new string that contains the full path of the resource being accessed.
    """
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)


# insert a image into the Excel sheet of three squares where the stamps of the teachers will go.
# todo add try except clause in case this cannot find the image, so that the program doesnt crash.
def add_shapes(new_sheet: Worksheet) -> None:
    """
    A function that retrieves an image in the same directory as this script and pastes it in the desired location.  This
    is an image of three boxes so that stamps of the person doing the daily meal reviews can be placed.
    :param new_sheet: The new sheet created where the image will be placed.
    :return: None
    """
    path = resource_path('image_boxes.jpg')
    image = Image(path)
    image.anchor = 'F1'
    new_sheet.add_image(image)


# insert the big kids meal data into a single Excel sheet. As well as the date and day of the week.
def insert_data_big_kids(date: int, data: tuple[str], new_sheet: Worksheet) -> None:
    """
    Access the desired cells and insert the meal data of the big kids from the tuple passed in.
    :param date: The date that this meal will be served.
    :param data: The tuple containing the meal data for this particular day that was passed in from a dictionary through
    iteration. here is the basic structure. (day of the week, breakfast, lunch, snack) and here is an example.
    {1: ('金', '●/▲鉄ウエハース\n【栗とかぼちゃとさつまいもクッキー】\nお茶', '鶏肉のカレームニエル\nさつま揚げとじゃが芋の味噌煮\nハムとわか
    めのサラダ\nごはん\n味噌汁', '●どら焼き(こしあん）\n【今川焼】\n▲牛乳【お茶】'), ...}
    :param new_sheet: The newly created sheet that the data will be inserted into.
    :return: None
    """
    day = data[0]
    breakfast = data[1]
    lunch = data[2]
    snack = data[3]
    new_sheet['B4'].value = new_sheet['B4'].value.replace('@', str(date))
    new_sheet['B4'].value = new_sheet['B4'].value.replace('$', day)
    new_sheet['C7'].value = breakfast
    new_sheet['C16'].value = lunch
    new_sheet['C25'].value = snack


# todo refactor code because some values are not need. perhaps change the structure of the dict as well.
# insert the small kids meal data into a single Excel sheet.
def insert_data_small_kids(date: int, data: tuple[str], new_sheet: Worksheet) -> None:
    """
    Access the desired cells and insert the meal data of the small kids from the tuple passed in.
    :param date: This is the date of when this particular meal will be served, but because the function that inserts the
    meal data for the big kids handles insertion of the date into the document this is not needed for this function.
    However, this variable will be left here in case refactoring is needed, and it becomes necessary.
    :param data: The tuple containing the meal data for this particular day that was passed in from a dictionary through
    iteration. here is the basic structure. (day of the week, breakfast, early, middle, late, snack) and here is an
    example. ('木', '野菜ハイハイン\n\n', '', '', '５倍粥\n鶏ササミと野菜（人参・グリンピース）煮物\n玉ねぎとわかめの煮物',
    'さつま芋きなこがけ\n\n') In this example there are no early and middle meals so they are just empty strings.
    The day of the week is not being used because again that is being handled by the other function that inserts the
    meal data for the big kids.
    :param new_sheet: The newly created sheet that the data will be inserted into.
    :return: None
    """
    breakfast = data[1]
    early = data[2]
    middle = data[3]
    late = data[4]
    snack = data[5]
    new_sheet['F7'].value = breakfast
    new_sheet['F16'].value = early
    new_sheet['F18'].value = middle
    new_sheet['F20'].value = late
    new_sheet['F25'].value = snack


# todo I fixed the issue when the user didn't pick a file because there is none for that month. I was able to handle the
# todo error but it raised another question. what if there is mismatching dates on two of the files. which would lead
# todo the key not being in the second dict which would mean it wont be pasted. This can happen if the person creating
# todo the files accidentally inserts the wrong days.  i might have to refactor code like this but it feels messy.
'''
if meal_data_small_kids is None:
    iterate with only meal_data_big_kids
else:
    iterate by zipping small and big kids meal data
'''


# insert the collected data for the big and small kids into each new Excel sheet.
def paste_meal_data(path: str, meal_data_big_kids: dict, meal_data_small_kids: dict) -> None:
    """
    Insert the data from the meal schedules for the big and small kids into the document that is used to rate the meals
    for the day.(for safety purposes.)
    :param path: The file path for the base document that will be used to create the new document.
    :param meal_data_big_kids: A dictionary that contains the data of the meal schedule for the big kids. The format of
    the dictionary is as follows {date: (day of the week, breakfast, lunch, snack), ...} and here is an example of the
    dictionary {4: ('木', '●/▲たまごボーロ\n【リッツクラッカー】\nお茶', '新春ちらし寿司\n切干大根の煮物\n花麩のすまし汁',
     '上用まんじゅう\n▲牛乳【お茶】'), ...}
    :param meal_data_small_kids: A dictionary that contains the data of the meal schedule for the small kids. The data
    structure for the small kids is slightly different with the lunch split into three depending on the development
    stage of the child. (early middle late) Also depending on the month of the year early and or middle will not be
    included, so they will be empty strings. Here is the basic format of the dictionary.
    date: (day of the week, breakfast, early, middle, late, snack), ...} and here is an example of the dictionary.
    {28: ('木', '野菜ハイハイン\n\n', '', '', '５倍粥\n鶏ササミと野菜（人参・グリンピース）煮物\n玉ねぎとわかめの煮物',
    'さつま芋きなこがけ\n\n'), ...} In this example the early and middle meals are not there so they are empty strings.
    :return: None
    """
    book = openpyxl.load_workbook(path)
    sheet = book.active
    for key, val_big in meal_data_big_kids.items():
        new_sheet = book.create_sheet(f'{key}({val_big[0]})')
        copy_all_elements(sheet, new_sheet)
        add_shapes(new_sheet)
        insert_data_big_kids(key, val_big, new_sheet)
        # checks to see if meal data is None and if the date exists in its dictionary.  meal data may be None if the
        # user did not choose a file to extract the data. dates may not exist if the Excel sheet contains errors that
        # make the dates differ between the big and small kids.
        if meal_data_small_kids and key in meal_data_small_kids:
            val_small = meal_data_small_kids[key]
            insert_data_small_kids(key, val_small, new_sheet)

    del book['base']  # remove base sheet because it will not be needed when printing.
    book.save(new_file_path(path, added_text='_test_complete'))


# the function that steps through the large steps of transferring the data.
def main() -> None:
    """
    The main function of the script. First prompts the user to choose the file path for the three Excel documents needed
    to create the output documents. Extracts data from the meal schedule for the big kids and then the small kids
    (the babies) and then pastes this information into the new document this script produces.
    """
    big_kids_path = choose_file(1)
    small_kids_path = choose_file(2)
    output_path = choose_file(3)
    meal_data_big_kids = extract_meal_data_big_kids(big_kids_path)
    meal_data_small_kids = extract_meal_data_small_kids(small_kids_path)
    print(meal_data_big_kids, meal_data_small_kids)
    paste_meal_data(output_path, meal_data_big_kids, meal_data_small_kids)


if __name__ == '__main__':
    main()
